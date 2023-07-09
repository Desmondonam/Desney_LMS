import tkinter as tk
import tkinter.messagebox as messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def login():
    username = username_entry.get()
    password = password_entry.get()

    # Perform your login validation here
    # Example: Check if username and password match certain criteria
    if username == "admin" and password == "password":
        messagebox.showinfo("Login Successful", "Welcome, " + username + "!")
        main_window()
    else:
        messagebox.showerror("Login Failed", "Invalid username or password")


def main_window():
    main = tk.Toplevel()
    main.title("Data Entry")

    # Create labels and entry fields for data entry
    name_label = tk.Label(main, text="Name:")
    name_label.grid(row=0, column=0, sticky="E")
    name_entry = tk.Entry(main)
    name_entry.grid(row=0, column=1)

    course_label = tk.Label(main, text="Course:")
    course_label.grid(row=1, column=0, sticky="E")
    course_entry = tk.Entry(main)
    course_entry.grid(row=1, column=1)

    subject_label = tk.Label(main, text="Subject:")
    subject_label.grid(row=2, column=0, sticky="E")
    subject_entry = tk.Entry(main)
    subject_entry.grid(row=2, column=1)

    marks_label = tk.Label(main, text="Marks:")
    marks_label.grid(row=3, column=0, sticky="E")
    marks_entry = tk.Entry(main)
    marks_entry.grid(row=3, column=1)

    submit_button = tk.Button(main, text="Submit", command=lambda: submit_data(
        name_entry, course_entry, subject_entry, marks_entry))
    submit_button.grid(row=4, columnspan=2)


def submit_data(name_entry, course_entry, subject_entry, marks_entry):
    name = name_entry.get()
    course = course_entry.get()
    subject = subject_entry.get()
    marks = marks_entry.get()

    # Load existing workbook or create a new one
    try:
        workbook = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save('data.xlsx')

    # Select the active sheet
    sheet = workbook.active

    # Check if the headers exist and create them if necessary
    headers = ["Name", "Course", "Subject", "Marks"]
    for col_num, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        if sheet[col_letter + '1'].value != header:
            sheet[col_letter + '1'] = header

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Write the data to the worksheet
    sheet.cell(row=next_row, column=1).value = name
    sheet.cell(row=next_row, column=2).value = course
    sheet.cell(row=next_row, column=3).value = subject
    sheet.cell(row=next_row, column=4).value = marks

    # Save the workbook
    workbook.save('data.xlsx')
    messagebox.showinfo("Success", "Data submitted successfully!")

    # Clear the input fields
    name_entry.delete(0, tk.END)
    course_entry.delete(0, tk.END)
    subject_entry.delete(0, tk.END)
    marks_entry.delete(0, tk.END)


# Create the login window
root = tk.Tk()
root.title("Login")

# Create labels and entry fields for username and password
username_label = tk.Label(root, text="Username:")
username_label.pack()
username_entry = tk.Entry(root)
username_entry.pack()

password_label = tk.Label(root, text="Password:")
password_label.pack()
password_entry = tk.Entry(root, show="*")
password_entry.pack()

# Create a login button
login_button = tk.Button(root, text="Login", command=login)
login_button.pack()

# Run the GUI
root.mainloop()
